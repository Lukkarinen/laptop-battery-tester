from openpyxl import Workbook
from datetime import datetime
import pathlib
import random

class Logging:
    def __init__(self, timestamp_format = "%Y-%m-%d, %H:%M:%S"):
        self.timestamp_format = timestamp_format

    def format_timestamp(self, timestamp):
        #Format a time object into a single string that's more suitable for writing into a dumb database or printing into terminal
        return timestamp.strftime(self.timestamp_format)

class ExcelFile(Logging):
    def __init__(self,
                 data_sheet_title = "Data",
                 date_column = "Date and time",
                 battery_column = "Battery charge (%)",
                 charger_column = "Charging",
                 log_folder = "\logs\\"
                 ):

        date = datetime.now().strftime("%Y-%m-%d-%H-%M")
        running_path = str(pathlib.Path(__file__).parent.resolve())
        self.log_folder = running_path + log_folder
        self.file_name = self.set_file_name(date, self.log_folder)

        self.workbook = Workbook()
        self.data_sheet = self.workbook.active
        self.data_sheet.title = data_sheet_title
        self.write_to_datasheet(self.data_sheet.max_row, date_column, battery_column, charger_column)

        super().__init__()

    def write_to_datasheet(self, row_number, date_column, battery_column, charge_column):
        if isinstance(date_column, datetime):
            date_column = self.format_timestamp(date_column)
        data_table = [date_column, battery_column, charge_column]
        for i in range (1,4):
            self.data_sheet.cell(row = row_number, column = i, value = data_table[i-1])

    def set_file_name(self, date, folder):
        for i in range(1,6):
            random_number = random.randint(100000, 999999)
            file_path = folder + date + "_%d" % random_number
            file_check = pathlib.Path(file_path)
            if not file_check.is_file():
                return file_path
        raise FileExistsError
    
    def next_row(self):
        return self.data_sheet.max_row + 1
    
    def save_file(self):
        self.workbook.save(self.file_name)

class Terminal(Logging):
    def __init__(self,
                 battery_error = "Error: Can't read battery state",
                 connect_charger = "Please plug the charger in",
                 disconnect_charger = "Please disconnect the charger",
                 stage_done = "Stage %s done",
                 ask_to_press_enter = "Press Enter to end the program..."
                 ):
        
        self.battery_error = battery_error
        self.connect_charger = connect_charger
        self.disconnect_charger = disconnect_charger
        self.stage_done = stage_done
        self.ask_to_press_enter = ask_to_press_enter

        super().__init__()
    
    def format_stage(self, stage):
        return self.stage_done % str(stage)
    
    def press_enter(self):
        input(self.ask_to_press_enter)
        
    def print_status(self, timestamp, event, stage=1):
        formatted_timestamp = self.format_timestamp(timestamp)

        match event:
            case "stage":
                message_text = self.format_stage(stage)
            case "disconnect":
                message_text = self.disconnect_charger
            case "connect":
                message_text = self.connect_charger
            case "error":
                message_text = self.battery_error
            case _:
                message_text = ""

        print("%s: %s." % (formatted_timestamp, message_text))